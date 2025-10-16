from openpyxl import load_workbook
from http.server import BaseHTTPRequestHandler
import json
import io
import urllib.request

TEMPLATE_URL = "https://raw.githubusercontent.com/1FAMM1/CB360-Mobile/main/templates/sitop_template.xlsx"

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            # Download do template
            req = urllib.request.Request(TEMPLATE_URL)
            with urllib.request.urlopen(req) as response:
                template_data = response.read()
            
            # Carrega o workbook (preserva form controls)
            wb = load_workbook(io.BytesIO(template_data), keep_vba=True)
            ws = wb.active
            
            # Preenche as células de texto
            ws['P11'] = data.get('vehicle', '')
            ws['E17'] = data.get('registration', '')
            ws['B17'] = data.get('gdh_inop', '')
            ws['O14'] = data.get('failure_type', '')
            
            desc = data.get('failure_description', '')
            ws['K16'] = f"Descrição: {desc}" if desc else ''
            
            ws['G23'] = data.get('gdh_op', '')
            ws['E28'] = data.get('optel', '')
            
            # Preenche células vinculadas às checkboxes (valores booleanos)
            ws['S1'] = bool(data.get('ppi_airport', False))
            ws['S2'] = bool(data.get('ppi_a22', False))
            ws['S3'] = bool(data.get('ppi_a2', False))
            ws['S4'] = bool(data.get('ppi_linfer', False))
            ws['S5'] = bool(data.get('ppi_airfield', False))
            ws['S6'] = bool(data.get('ppi_part', False))
            ws['S7'] = not bool(data.get('ppi_part', False))
            ws['S8'] = bool(data.get('ppi_subs', False))
            ws['S9'] = not bool(data.get('ppi_subs', False))
            
            # Salva em buffer
            output = io.BytesIO()
            wb.save(output)
            excel_data = output.getvalue()  # Pega os bytes diretamente
            
            # Nome do ficheiro
            vehicle = data.get('vehicle', 'vehicle')
            gdh = data.get('gdh_inop', '')
            filename = f"SITOP_{vehicle}_{gdh}.xlsx"
            
            # Retorna o ficheiro (IMPORTANTE: sem encoding, dados binários puros)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Length', str(len(excel_data)))
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Expose-Headers', 'Content-Disposition')
            self.end_headers()
            
            # Escreve os dados binários diretamente
            self.wfile.write(excel_data)
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            error_response = json.dumps({
                'error': str(e), 
                'type': type(e).__name__,
                'detail': error_detail
            })
            self.wfile.write(error_response.encode('utf-8'))
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
